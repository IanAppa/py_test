# https://www.coupang.com/np/search/autoComplete?callback=jQuery&keyword=%EC%96%91%EB%A7%90
import requests
import json
import os
import openpyxl
from bs4 import BeautifulSoup as bs

excelFile = os.getcwd() + '\keyword_list.xlsx'
wb = openpyxl.load_workbook(excelFile)
ws = wb.active

##for row in range(2, ws.max_row + 1) :
for row in range(2, 7) :
    baseurl = 'https://www.coupang.com/np/search/autoComplete?callback=jQuery&keyword='
    keywordStr = ws.cell(row, 9).value
    keyword = keywordStr.strip().split("/")
    col = 0
    for i in range(0, len(keyword)) :
        log = str(row) + '●' + keyword[i] + ' : '

        response = requests.get(baseurl + keyword[i])

        jsondata = response.text
        jsondata = jsondata[7:len(jsondata)-1]
        #print(data)

        data = json.loads(jsondata)

        #print(ws.cell(1, 1).value)
        for j in range(0, len(data)) :
            col += 1
            log += data[j]['keyword'] + ', '
            ws.cell(row, col + 9, data[j]['keyword'])
        print(log[0:len(log)-2])

    wb.save(excelFile)
    wb.close()
#r = response.json()
#print(response.json())